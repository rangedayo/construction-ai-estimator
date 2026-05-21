"""도면 정답지(xlsx) 로더 — 회귀 테스트용 결정론적 정답 데이터.

라운드 2 방침: 페이지 분할 폐기. 정답지를 페이지 단위로 분해하지 않고
'도면 전체에서 부호별 합계'만 로드한다.

정답지(`reference_materials/도면 정답지.xlsx`)는 시트마다 매트릭스 형태:
    행 = 도면 페이지,  열 = 부재 부호,  셀 = 해당 페이지의 부호 개수

셀 값 해석 규칙:
    빈 셀(None)  → 그 페이지에 그 부호 없음
    0            → 부호는 등장하나 개수 0개
    '합계' 행/열  → 집계용이므로 제외 (페이지 셀을 직접 합산한다)

LLM 호출 등 비결정 요소 없이 순수하게 xlsx만 파싱한다.
"""
from __future__ import annotations

import os

import openpyxl

# tests/ground_truth.py → poc_v2/tests → poc_v2 → 프로젝트 루트
_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
ANSWER_KEY_PATH = os.path.join(
    PROJECT_ROOT, "reference_materials", "도면_정답지.xlsx"
)
SYMBOL_RULES_PATH = os.path.join(
    PROJECT_ROOT, "config", "symbol_rules.yaml"
)

_TOTAL_LABEL = "합계"


def load_text_height_filter(
    path: str | None = None,
) -> dict[str, float | None]:
    """symbol_rules.yaml 의 text_height_filter → {도면명: min_height}.

    min_height 가 null(YAML) 이면 None 으로, 숫자면 그대로 돌려준다.
    설정 파일이나 pyyaml 이 없으면 빈 dict (= 모든 도면 필터 미적용).
    """
    config = _load_config(path)
    raw = config.get("text_height_filter", {}) or {}
    return {
        drawing: (spec or {}).get("min_height")
        for drawing, spec in raw.items()
    }


def load_auto_policy_params(
    path: str | None = None,
) -> dict[str, float]:
    """symbol_rules.yaml 의 auto_policy_params → 자동 정책 판단 파라미터.

    라운드 6: 신호 3(규격 패턴) 임계값. 누락 시 auto_detect_policy 기본값.
    """
    config = _load_config(path)
    raw = config.get("auto_policy_params", {}) or {}
    return {
        "spec_pattern_threshold": float(raw.get("spec_pattern_threshold", 0.3)),
    }


def load_policy_override(
    drawing_name: str,
    path: str | None = None,
) -> dict[str, bool] | None:
    """symbol_rules.yaml 의 policy_override[drawing_name] 로드.

    라운드 6: 자동 판단을 강제로 덮어쓰는 비상용 설정. yaml 값이 null 이거나
    키가 없으면 None 을 돌려준다(= 자동 판단 사용). dict 면 두 플래그를 채운
    정규화된 dict 를 돌려준다.
    """
    config = _load_config(path)
    raw = (config.get("policy_override", {}) or {}).get(drawing_name)
    if not raw:
        return None
    return {
        "exclude_table_regions": bool(raw.get("exclude_table_regions", False)),
        "exclude_with_spec": bool(raw.get("exclude_with_spec", False)),
    }


def load_policy_p(
    path: str | None = None,
) -> dict[str, dict[str, bool]]:
    """[DEPRECATED — 라운드 6] symbol_rules.yaml 의 policy_p → {도면명: {플래그}}.

    라운드 5 정책 P 도면별 수동 분기 로더. 라운드 6 에서 policy_p 키는
    주석 처리되고 신호 2·3 은 auto_policy.auto_detect_policy 가 자동 판단한다.
    yaml 키가 폐기됐으므로 이 함수는 빈 dict 를 돌려준다. 호출 금지.
    """
    config = _load_config(path)
    raw = config.get("policy_p", {}) or {}
    result: dict[str, dict[str, bool]] = {}
    for drawing, spec in raw.items():
        spec = spec or {}
        result[drawing] = {
            "exclude_table_regions": bool(spec.get("exclude_table_regions", False)),
            "exclude_with_spec": bool(spec.get("exclude_with_spec", False)),
        }
    return result


def load_table_region_params(
    path: str | None = None,
) -> dict[str, float | int]:
    """symbol_rules.yaml 의 table_region_detection → 일람표 검출 파라미터.

    누락 시 detect_table_regions 기본값과 같은 값으로 채운다.
    """
    config = _load_config(path)
    raw = config.get("table_region_detection", {}) or {}
    return {
        "region_size_ratio": float(raw.get("region_size_ratio", 1 / 30)),
        "min_distinct_symbols": int(raw.get("min_distinct_symbols", 4)),
        "max_count_per_symbol": int(raw.get("max_count_per_symbol", 2)),
    }


def _load_config(path: str | None) -> dict:
    """symbol_rules.yaml 파싱. pyyaml·파일 부재 시 빈 dict."""
    try:
        import yaml  # noqa: PLC0415
    except ImportError:
        return {}
    cfg_path = path or SYMBOL_RULES_PATH
    if not os.path.exists(cfg_path):
        return {}
    with open(cfg_path, encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def drawing_symbol_totals(
    path: str | None = None,
) -> dict[str, dict[str, int]]:
    """{도면명: {부호: 도면 전체 합계}} 형태로 정답지를 파싱해 반환한다.

    정답지의 '합계' 행을 신뢰하지 않고 페이지 셀에서 직접 합산한다
    (합계 행은 수기 입력이라 오차가 있을 수 있음). 빈 셀은 0으로 보고
    한 번이라도 등장한 부호만 결과에 포함한다.
    """
    workbook = openpyxl.load_workbook(path or ANSWER_KEY_PATH, data_only=True)
    totals: dict[str, dict[str, int]] = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue

        header = rows[0]
        # (열 인덱스, 부호명) — 0번 열(도면명)과 '합계' 열은 제외
        symbol_columns: list[tuple[int, str]] = []
        for col_idx, raw_name in enumerate(header[1:], start=1):
            if raw_name is None:
                continue
            label = str(raw_name).strip()
            if not label or label == _TOTAL_LABEL:
                continue
            symbol_columns.append((col_idx, label))

        aggregated: dict[str, int] = {}
        for row in rows[1:]:
            raw_page = row[0]
            if raw_page is None:
                continue
            page_name = str(raw_page).strip()
            if not page_name or page_name == _TOTAL_LABEL:
                continue
            for col_idx, symbol in symbol_columns:
                value = row[col_idx] if col_idx < len(row) else None
                if value is None or value == "":
                    continue
                aggregated[symbol] = aggregated.get(symbol, 0) + int(value)

        totals[sheet_name] = aggregated

    return totals


def within_tolerance(
    predicted: int,
    expected: int,
    rel_tol: float = 0.05,
    small_count: int = 5,
) -> bool:
    """예측 개수가 정답 허용 오차 안에 있는지 판정.

    expected 가 small_count(기본 5) 이하면 ±1 까지 허용한다(작은 수에서
    5%는 너무 빡빡함). 그 외에는 상대오차 rel_tol(기본 5%) 이하면 통과.
    """
    diff = abs(predicted - expected)
    if expected <= small_count:
        return diff <= 1
    return diff <= expected * rel_tol
