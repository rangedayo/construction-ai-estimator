"""정답지 시트명 매칭 — 라운드 베이스라인-2 작업 2.

표제부에서 추출한 도면명(list[str])을 정답지 시트명·길이 라우팅 라벨과
정규화 후 결정론적으로 매칭한다.

매칭 순서 (보편 룰 우선; fallback yaml 은 최후의 수단)
    1. exact    : 정규화 일치 (카운트 시트 우선, 그다음 길이 라벨)
    2. partial  : 한쪽이 다른쪽을 포함 (도면4 "단면도" ⊂ "종단면도"/"횡단면도")
    3. fallback : config/sheet_name_overrides.yaml 의 매핑
    4. unmatched: 위 셋 실패

데이터 소스 (무수정)
    * reference_materials/도면_정답지.xlsx — 카운트 시트(세부 도면명 = 데이터 행)
    * config/length_routing.yaml          — 길이 측정 소스 라벨(sheet_label)
"""
from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from typing import Literal, Optional

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

_COUNT_ANSWER_KEY = os.path.join(
    PROJECT_ROOT, "reference_materials", "도면_정답지.xlsx"
)
_LENGTH_ROUTING = os.path.join(PROJECT_ROOT, "config", "length_routing.yaml")
_OVERRIDES_PATH = os.path.join(PROJECT_ROOT, "config", "sheet_name_overrides.yaml")

_TOTAL_LABEL = "합계"
_NON_SYMBOL_HEADERS = frozenset({"도면명", "분석 대상", _TOTAL_LABEL})
_CATEGORY_SUFFIXES = ("-기둥", "-보")

Confidence = Literal["exact", "partial", "fallback", "unmatched"]
Kind = Literal["count", "length", "unmatched"]

# 정규화 제거 문자 — 공백·콤마·괄호·줄바꿈·하이픈·슬래시·점·중점.
_NORMALIZE_STRIP = re.compile(r"[\s,()\[\]\-/.·、]+")


@dataclass(frozen=True)
class SheetMatch:
    """추출 도면명 ↔ 정답지 시트 매칭 결과."""
    matched_sheet: Optional[str]
    confidence: Confidence
    kind: Kind
    candidates: list[str]


def normalize(text: str) -> str:
    """공백·구분자 제거 + 소문자화한 매칭 키."""
    return _NORMALIZE_STRIP.sub("", text).lower()


def _drawing_name_from_sheet(sheet_name: str) -> tuple[str, Optional[str]]:
    for suffix in _CATEGORY_SUFFIXES:
        if sheet_name.endswith(suffix):
            return sheet_name[: -len(suffix)], suffix[1:]
    return sheet_name, None


def _parse_cell_int(value: object) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def load_count_sheet_rows(
    drawing: str,
    path: Optional[str] = None,
    category: Optional[str] = None,
) -> dict[str, dict[str, int]]:
    """정답지에서 {세부도면명: {부호: 개수}} 를 행 단위로 로드.

    `drawing_symbol_totals` 는 행을 합산해 시트명을 잃지만, 베이스라인-2 는
    시트(세부 도면)별 정답이 필요하다. 합계 행은 건너뛴다.

    category=None 이면 기둥·보 시트를 병합(매칭 후보용), "기둥"/"보" 면 해당
    카테고리 시트만 로드(카운트 PASS 비교용 — 이번 라운드는 기둥 스코프).
    """
    workbook = openpyxl.load_workbook(path or _COUNT_ANSWER_KEY, data_only=True)
    rows_by_sheet: dict[str, dict[str, int]] = {}

    for sheet_name in workbook.sheetnames:
        drawing_name, sheet_category = _drawing_name_from_sheet(sheet_name)
        if drawing_name != drawing:
            continue
        if category is not None and sheet_category != category:
            continue
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        symbol_columns: list[tuple[int, str]] = []
        for col_idx, raw_name in enumerate(header):
            if raw_name is None:
                continue
            label = str(raw_name).strip()
            if not label or label in _NON_SYMBOL_HEADERS:
                continue
            symbol_columns.append((col_idx, label))

        for row in rows[1:]:
            if not row:
                continue
            raw_first = row[0]
            sheet_label = "" if raw_first is None else str(raw_first).strip()
            if not sheet_label or sheet_label == _TOTAL_LABEL:
                continue
            bucket = rows_by_sheet.setdefault(sheet_label, {})
            for col_idx, symbol in symbol_columns:
                value = row[col_idx] if col_idx < len(row) else None
                count = _parse_cell_int(value)
                if count <= 0:
                    continue
                bucket[symbol] = bucket.get(symbol, 0) + count
    return rows_by_sheet


def load_length_labels(
    drawing: str,
    path: Optional[str] = None,
) -> list[str]:
    """length_routing.yaml 의 sheet_label 을 라벨 토큰 리스트로 분해.

    예) 도면4 "종단면도, 횡단면도" → ["종단면도", "횡단면도"].
    """
    try:
        import yaml  # noqa: PLC0415
    except ImportError:
        return []
    cfg_path = path or _LENGTH_ROUTING
    if not os.path.exists(cfg_path):
        return []
    with open(cfg_path, encoding="utf-8") as handle:
        config = yaml.safe_load(handle) or {}
    entry = (config.get("drawings", {}) or {}).get(drawing)
    if not entry:
        return []
    labels: list[str] = []
    for source in entry.get("sources", []) or []:
        raw = source.get("sheet_label", "")
        # 리스트 구분자는 "콤마+공백"(", ") 또는 "、" 만 인정한다. 콤마가 토큰
        # 내부("가,나동" = 가동·나동 결합)에 쓰일 수 있으므로 공백 없는 콤마로는
        # 쪼개지 않는다. 예) "종단면도, 횡단면도" → 2개,  "가,나동 횡단면도" → 1개.
        for token in re.split(r",\s+|\s*、\s*", str(raw)):
            token = token.strip()
            if token and token not in labels:
                labels.append(token)
    return labels


def load_overrides(
    drawing: str,
    path: Optional[str] = None,
) -> dict[str, list[str]]:
    """sheet_name_overrides.yaml 의 {drawing: {표제부텍스트: [타깃]}} 로드.

    파일·pyyaml 부재 시 빈 dict (= override 없음).
    """
    try:
        import yaml  # noqa: PLC0415
    except ImportError:
        return {}
    cfg_path = path or _OVERRIDES_PATH
    if not os.path.exists(cfg_path):
        return {}
    with open(cfg_path, encoding="utf-8") as handle:
        config = yaml.safe_load(handle) or {}
    raw = (config.get(drawing, {}) or {})
    result: dict[str, list[str]] = {}
    for key, targets in raw.items():
        if isinstance(targets, str):
            result[str(key)] = [targets]
        elif isinstance(targets, list):
            result[str(key)] = [str(t) for t in targets]
    return result


def _find_exact(norm_titles: set[str], norm_targets: dict[str, str]) -> Optional[str]:
    for norm, raw in norm_targets.items():
        if norm in norm_titles:
            return raw
    return None


def _find_partial(norm_titles: set[str], norm_targets: dict[str, str]) -> Optional[str]:
    for norm, raw in norm_targets.items():
        for title in norm_titles:
            if not title or not norm:
                continue
            if title in norm or norm in title:
                return raw
    return None


def match_sheet(
    drawing: str,
    extracted_titles: list[str],
    *,
    count_path: Optional[str] = None,
    length_path: Optional[str] = None,
    overrides_path: Optional[str] = None,
) -> SheetMatch:
    """추출 도면명 후보를 정답지 시트·길이 라벨과 매칭한다."""
    count_rows = load_count_sheet_rows(drawing, count_path)
    length_labels = load_length_labels(drawing, length_path)

    norm_titles = {normalize(t) for t in extracted_titles if t}
    norm_count = {normalize(name): name for name in count_rows}
    norm_length = {normalize(lbl): lbl for lbl in length_labels}
    candidates = list(count_rows.keys()) + length_labels

    if not norm_titles:
        return SheetMatch(None, "unmatched", "unmatched", candidates)

    # 1) exact — 실제 카운트 대상(비어있지 않은 행)이 최우선.
    #    단, 카운트 행이 placeholder(빈 dict = 부호 0개)이고 동시에 length
    #    라벨(length_routing 이 명시한 측정 소스)이면 length 가 우선한다.
    #    예) 도면5 주단면도1·4 는 카운트 시트에 0-행으로 올라와 있지만 실제로는
    #    단면도(길이 측정 소스)다. length_routing 은 사람이 지정한 신호이므로
    #    auto-populated placeholder 행보다 강하다.
    hit_count = _find_exact(norm_titles, norm_count)
    hit_length = _find_exact(norm_titles, norm_length)
    if hit_count is not None:
        if hit_length is not None and not count_rows.get(hit_count):
            return SheetMatch(hit_length, "exact", "length", candidates)
        return SheetMatch(hit_count, "exact", "count", candidates)
    if hit_length is not None:
        return SheetMatch(hit_length, "exact", "length", candidates)

    # 2) partial
    hit = _find_partial(norm_titles, norm_count)
    if hit is not None:
        return SheetMatch(hit, "partial", "count", candidates)
    hit = _find_partial(norm_titles, norm_length)
    if hit is not None:
        return SheetMatch(hit, "partial", "length", candidates)

    # 3) fallback yaml
    overrides = load_overrides(drawing, overrides_path)
    for title in extracted_titles:
        targets = overrides.get(title)
        if targets:
            kind: Kind = "count" if targets[0] in count_rows else "length"
            return SheetMatch(targets[0], "fallback", kind, candidates)

    # 4) unmatched
    return SheetMatch(None, "unmatched", "unmatched", candidates)
